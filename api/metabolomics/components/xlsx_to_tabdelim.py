#! /usr/bin/env python
# -*- coding: utf-8 -*-
__author__ = 'Matthew L. Bendall'

import re
import sys
import argparse
from openpyxl import load_workbook

def convert_to_tab_delim(source_path, destination_path):

    f = open(source_path, 'rb')
    wb = load_workbook(f)

    #ws = wb[wb.get_sheet_names()[args.sheetnum-1]]
    ws = wb[wb.sheetnames[0]]
    totalrows = 0
    with open(destination_path, "w") as f:
        for i,row in enumerate(ws.rows):
            vals = []
            for cell in row:
                if cell.value is None:
                    val = ''
                else:
                    val = str(cell.value)
                    val = re.sub(r'\s', ' ', val)
                    val = re.sub(r"'", '', val)
                vals.append(val)

            print ('\t'.join(vals),  file=f)
            totalrows = i + 1
